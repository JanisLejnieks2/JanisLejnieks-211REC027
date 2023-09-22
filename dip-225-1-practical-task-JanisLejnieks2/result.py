from openpyxl import Workbook, load_workbook 
wb=load_workbook('tests/test1.xlsx')
ws=wb.active
total=0
#write your code here
max_row=ws.max_row
for row in range(2, max_row+1):
    rate=ws['B'+str(row)].value
    hours=ws['C'+str(row)].value
   
    if (type(hours)!=str and type(rate)!=str):
        salary=(hours)*(rate)
        ws['D'+str(row)].value=salary
        if (salary>3000):
            total = total+1
print(total)
wb.save('tests/testComplete.xlsx')
wb.close()


    